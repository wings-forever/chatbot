from selenium import webdriver
import openpyxl

driver=webdriver.Chrome("*/chromedriver")
driver.implicitly_wait(1)


#go to the login page & login

driver.get("https://everytime.kr/login")
driver.find_element_by_name("userid").send_keys("*")
driver.find_element_by_name("password").send_keys("*")
driver.find_element_by_xpath('//*[@class="submit"]/input').click()
driver.implicitly_wait(1)

#본문,댓글, 작성자명 리스트
main_results=[]
comment_results=[]
poster_name=[]

#본문,댓글 링크 리스트
main_urls=[]
comment_urls=[]

cnt=0
while True:
    print("Page "+str(cnt))

    if cnt > 8:   #163페이지까지가 2019년도 게시글
        break
    cnt=cnt+1

    driver.get("https://everytime.kr/375135/p/"+str(cnt))
    driver.implicitly_wait(1)

    #get articles link
    posts=driver.find_elements_by_css_selector("article > a.article")
    links=[post.get_attribute("href") for post in posts]



   #get detail article
    for link in links:
        driver.get(link)

        posters = driver.find_elements_by_css_selector("h3.large") #작성자

        main_articles = driver.find_elements_by_css_selector("p.large")
        comments = driver.find_elements_by_css_selector("div.comments > article.parent > p.large")  # 댓글(가장 상위에 있는)
        child_comments = driver.find_elements_by_css_selector("div.comments > article.child > p.large") #대댓글

        for poster in posters:
            poster_name.append(poster.text)

        sweet_vanila="달달한바닐라" #'달달한바닐라'님 글만 크롤링

        if sweet_vanila in poster_name:
            for main_article in main_articles:
                main_results.append(main_article.text)   #main_results -> 본문+댓글 리스트
                main_urls.append(link)

            for comment in comments:
                comment_results.append(comment.text)     #comment_results -> 댓글 리스트
                comment_urls.append(link)

            for child_comment in child_comments:
                comment_results.append(child_comment.text) #대댓글도 comment_results에 추가
                comment_urls.append(link)

#print(main_results)
#print(comment_results)

wb = openpyxl.Workbook()
sheet1 = wb.create_sheet('sweet_vanila', 1)

for i in range(len(main_results)):
    sheet1.cell(row=i+1, column=1).value = main_results[i]
    sheet1.cell(row=i+1, column=2).value = main_urls[i]

wb.save('sweet_vanila_results.xlsx')
